"""Export the patient database to a real Excel ``.xlsx`` workbook (openpyxl).

One sheet, one row per scan/visit, sorted by patient (name then id) then acquisition date — so each
patient's visits read chronologically. The header is bold and frozen; the GA-area column is a real
number (so Excel can chart/sum it), the date is kept as the device's display string (Spectralis date
strings are not always parseable, and the pre-sort already orders the rows).

openpyxl is a pure-python dependency added for the clinic (the only new one). It is imported lazily so
this module — and the rest of the app — still imports if the package is somehow absent; the route maps
that to a clear 'export unavailable' error rather than a crash.
"""
import io

# Column spec: (header, db-field, kind). kind 'num' => numeric cell with a format; else text.
_COLUMNS = [
    ("Patient ID", "patient_id", "text"),
    ("Patient name", "patient_name", "text"),
    ("Eye", "eye", "text"),
    ("Date", "_date", "text"),
    ("GA area (mm²)", "ga_area_mm2", "num"),
    ("BM source", "bm_source", "text"),
    ("B-scans", "n_bscans", "int"),
    ("FOV (mm)", "_fov", "text"),
]


def _date(row):
    return (row.get("acq_date") or "").split(" ")[0] or (row.get("logged_at") or "").split("T")[0] or ""


def _fov(row):
    w, h = row.get("fov_w_mm"), row.get("fov_h_mm")
    return f"{w}×{h}" if (w and h) else ""


def _sorted(rows):
    return sorted(
        rows,
        key=lambda r: (((r.get("patient_name") or "~").lower(), r.get("patient_id") or ""),
                       r.get("acq_iso") or r.get("record_id") or "", r.get("eye") or ""),
    )


def workbook_bytes(rows) -> bytes:
    """Build the ``.xlsx`` from already-read (enriched) DB rows; return the file bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError as e:                                   # pragma: no cover (deps issue, not data)
        raise RuntimeError("Excel export needs the 'openpyxl' package, which is not installed.") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "GA measurements"

    bold = Font(bold=True)
    for c, (header, _field, _kind) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = bold
        cell.alignment = Alignment(vertical="center")

    for r_i, row in enumerate(_sorted(rows), start=2):
        for c, (_header, field, kind) in enumerate(_COLUMNS, start=1):
            if field == "_date":
                value = _date(row)
            elif field == "_fov":
                value = _fov(row)
            else:
                value = row.get(field)
            cell = ws.cell(row=r_i, column=c)
            if kind == "num":
                try:
                    cell.value = float(value)
                    cell.number_format = "0.0000"
                except (TypeError, ValueError):
                    cell.value = "" if value is None else value
            elif kind == "int":
                try:
                    cell.value = int(value)
                except (TypeError, ValueError):
                    cell.value = "" if value is None else value
            else:
                cell.value = "" if value is None else str(value)

    # Freeze the header row and set readable column widths.
    ws.freeze_panes = "A2"
    widths = [14, 22, 6, 10, 14, 14, 12, 9, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
